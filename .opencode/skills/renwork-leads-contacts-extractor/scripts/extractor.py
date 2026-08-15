"""
RenWork (OKKI) RPA Leads Contacts Extractor CLI
- Dynamically loads company names from any Excel file and column.
- Automatically handles both Electron windows and embedded React drawer layouts.
- Safely writes extracted contacts to the "17. 采购决策人联系方式" sheet in the same workbook.
- Supports CLI arguments for spreadsheet path, sheet name, column name, and pagination limits.
"""

import asyncio
import json
import os
import re
import sys
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import argparse

try:
    from pyppeteer import connect
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "pyppeteer", "-q"])
    from pyppeteer import connect

EMAIL_REGEX = re.compile(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}')

def log(msg):
    ts = time.strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)

async def search_company(page, company):
    """搜索目标公司 (使用 Puppeteer 原生 click/type/press(Enter) 以规避 React 状态更新延迟)"""
    inp_selector = 'input[placeholder*="营销产品"]'
    try:
        # 1. 等待输入框可见并点击以获得焦点
        await page.waitForSelector(inp_selector, {'timeout': 5000})
        await page.click(inp_selector)
        
        # 2. 全选文本并删除以清空内容
        await page.keyboard.down('Meta')
        await page.keyboard.press('KeyA')
        await page.keyboard.up('Meta')
        await page.keyboard.press('Backspace')
        await asyncio.sleep(0.2)
        
        # 3. 模拟键盘逐字输入公司名
        await page.type(inp_selector, company, {'delay': 10})
        await asyncio.sleep(0.5)
        
        # 4. 原生按下回车触发搜索
        await page.keyboard.press('Enter')
        return {"ok": True, "method": "Native Click+Type+Enter"}
    except Exception as e:
        log(f"  [RPA 警告] 原生搜索输入失败: {e}，启用 JS 属性设置器兜底...")
        return await page.evaluate("""async (kw) => {
            const inputs = Array.from(document.querySelectorAll('input'));
            const inp = inputs.find(i => i.placeholder && i.placeholder.includes('营销产品') && i.offsetWidth > 0);
            if (!inp) return { ok: false, reason: '找不到搜索输入框' };
            const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
            setter.call(inp, kw);
            inp.dispatchEvent(new Event('input', { bubbles: true }));
            inp.dispatchEvent(new Event('change', { bubbles: true }));
            await new Promise(r => setTimeout(r, 600));
            inp.dispatchEvent(new KeyboardEvent('keydown', { key: 'Enter', keyCode: 13, bubbles: true }));
            return { ok: true, method: 'JS Fallback' };
        }""", company)

async def click_emails_link(page):
    """Click email count link of first row in search results"""
    return await page.evaluate("""() => {
        const els = Array.from(document.querySelectorAll('.contact-text, .contact-count-num, span'));
        const mailEl = els.find(el => el.textContent.trim() === '邮箱' && el.offsetWidth > 0);
        if (mailEl) {
            mailEl.click();
            return { ok: true, trigger: '邮箱文本' };
        }
        const countEl = els.find(el => el.className.includes('contact-count-num') && el.offsetWidth > 0);
        if (countEl) {
            countEl.click();
            return { ok: true, trigger: '邮箱数量' };
        }
        return { ok: false };
    }""")

async def scrape_contacts_from_page(page, company_name, max_pages=3):
    """Scrape paginated contacts from details view (either window or drawer)"""
    company_contacts = []
    current_page = 1
    
    while current_page <= max_pages:
        log(f"  → 正在抓取第 {current_page} 页联系人列表...")
        contacts_list = await page.evaluate("""() => {
            const rows = Array.from(document.querySelectorAll('tr'));
            return rows.filter(r => r.className === 'okki-table-row okki-table-row-level-0' && r.offsetWidth > 0)
                .map(r => {
                    const tds = Array.from(r.querySelectorAll('td')).map(td => td.textContent.trim());
                    return {
                        name: tds[2] || '',
                        title: tds[3] || '',
                        raw_email: tds[4] || '',
                        raw_phone: tds[5] || '',
                        score: tds[7] || '',
                        date: tds[9] || ''
                    };
                });
        }""")
        log(f"  → 第 {current_page} 页读取到 {len(contacts_list)} 个联系人")
        
        for c in contacts_list:
            emails = EMAIL_REGEX.findall(c["raw_email"])
            phones = [p.strip() for p in c["raw_phone"].split('+') if p.strip()]
            cleaned_phones = []
            for p in phones:
                if not p.startswith('+'):
                    p = '+' + p
                cleaned_phones.append(p)
                
            company_contacts.append({
                "company": company_name,
                "name": c["name"],
                "title": c["title"],
                "emails": emails,
                "phones": cleaned_phones,
                "score": c["score"],
                "date": c["date"]
            })
            
        next_page_str = str(current_page + 1)
        has_next = await page.evaluate("""async (next_str) => {
            const els = Array.from(document.querySelectorAll('a, li, span, button'));
            const nextBtn = els.find(el => el.textContent.trim() === next_str && el.offsetWidth > 0 && el.offsetHeight > 0);
            if (nextBtn) {
                nextBtn.click();
                return true;
            }
            return false;
        }""", next_page_str)
        
        if has_next:
            log(f"  → 点击翻页至第 {next_page_str} 页...")
            await asyncio.sleep(3.5)
            current_page += 1
        else:
            log(f"  → 无更多联系人页面")
            break
            
    return company_contacts

async def run_leads_contacts_extraction_client(excel_path, sheet_name, column_name, max_pages):
    if not os.path.exists(excel_path):
        log(f"错误: Excel文件不存在: {excel_path}")
        return
        
    log(f"正在加载 Excel 文件: {excel_path} ...")
    wb = openpyxl.load_workbook(excel_path)
    
    # Auto-detect sheet if not specified
    if not sheet_name:
        for name in wb.sheetnames:
            if "主候选" in name or "5." in name or "公司" in name or "Leads" in name:
                sheet_name = name
                break
        if not sheet_name:
            sheet_name = wb.sheetnames[0]
            
    ws = wb[sheet_name]
    log(f"已选定工作表: '{sheet_name}'")
    
    # Find company name column
    headers = [cell.value for cell in ws[1]]
    col_idx = None
    if column_name:
        if column_name in headers:
            col_idx = headers.index(column_name) + 1
    else:
        for opt in ["公司名称", "公司名称 (Company Name)", "Company Name", "真实买方公司", "买家公司名称"]:
            if opt in headers:
                col_idx = headers.index(opt) + 1
                column_name = opt
                break
        if not col_idx:
            # Fallback to column A
            col_idx = 1
            column_name = headers[0] if len(headers) > 0 else "Column A"
            
    log(f"已选定公司名称列: '{column_name}' (列号: {col_idx})")
    
    companies = []
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=col_idx).value
        if val:
            companies.append(str(val).strip())
            
    log(f"共加载 {len(companies)} 家公司。")
    if len(companies) == 0:
        log("无公司需要抓取。")
        return
        
    try:
        browser = await connect(browserURL='http://127.0.0.1:9222')
        log("[RPA] CDP 连接成功！")
    except Exception as e:
        log(f"[RPA 错误] 无法连接到客户端 (请确保 OKKI 客户端已开启，并配置了 CDP 远程端口 9222): {e}")
        return
        
    pages = await browser.pages()
    main_page = None
    for p in pages:
        title = await p.title()
        if "智能获客" in title or "mining-v2" in p.url:
            main_page = p
            break
    if not main_page:
        for p in pages:
            if "shell/index.html" in p.url:
                main_page = p
                break
    if not main_page:
        main_page = pages[0]
        
    log(f"[RPA] 锁定主搜索页面: {main_page.url}")
    
    await main_page.evaluate("() => { window.location.hash = '#/new_discovery/mining-v2/list'; }")
    await asyncio.sleep(4)
    
    # Clean old popups/drawers
    pages = await browser.pages()
    for p in pages:
        if "crm.xiaoman.cn/new_discovery/company" in p.url:
            log(f"[RPA] 关闭残留详情窗口 {p.url[:60]}...")
            await p.close()
            await asyncio.sleep(1)
            
    await main_page.evaluate("""() => {
        const mask = document.querySelector('.okki-drawer-mask');
        if (mask) mask.click();
    }""")
    await asyncio.sleep(1)
    
    all_contacts_database = []
    workspace_dir = os.path.dirname(os.path.abspath(excel_path))
    
    for idx, company in enumerate(companies):
        log(f"\n{'─'*55}")
        log(f"[{idx+1:03d}/{len(companies)}] 正在检索: 「{company}」")
        log(f"{'─'*55}")
        
        # 1. Search
        sr = await search_company(main_page, company)
        log(f"  → 搜索指令发送结果: {sr}")
        await asyncio.sleep(4)
        
        # 2. Click email link
        click_res = await click_emails_link(main_page)
        log(f"  → 点击邮箱链接结果: {click_res}")
        if not click_res["ok"]:
            log(f"  📋 [结果] ❌ 未在列表找到该公司的邮箱入口，可能无检索结果，跳过。")
            continue
            
        # 3. Wait and detect mode
        await asyncio.sleep(4)
        
        # Check window popup
        all_pages = await browser.pages()
        detail_page = None
        for p in all_pages:
            if "crm.xiaoman.cn/new_discovery/company" in p.url:
                detail_page = p
                break
                
        company_contacts = []
        if detail_page:
            d_title = await detail_page.title()
            log(f"  → [窗口模式] ✅ 成功捕获详情窗口: {d_title!r}")
            try:
                await detail_page.waitForSelector('.okki-tabs-tab-btn', {'timeout': 8000})
                tab_res = await detail_page.evaluate("""() => {
                    const btns = Array.from(document.querySelectorAll('.okki-tabs-tab-btn'));
                    const target = btns.find(b => b.textContent.trim().includes('联系人'));
                    if (target) {
                        target.click();
                        return { ok: true, text: target.textContent.trim() };
                    }
                    return { ok: false };
                }""")
                log(f"  → 切换详情页联系人 Tab 结果: {tab_res}")
                await asyncio.sleep(3)
                
                company_contacts = await scrape_contacts_from_page(detail_page, company, max_pages)
            except Exception as e:
                log(f"  ❌ 抓取窗口模式联系人出错: {e}")
            finally:
                log(f"  → 正在关闭详情窗口...")
                await detail_page.close()
                await asyncio.sleep(2)
        else:
            # Check drawer layout
            has_drawer = await main_page.evaluate("() => !!document.querySelector('.okki-drawer.okki-drawer-open')")
            if has_drawer:
                log(f"  → [抽屉模式] ✅ 检测到内嵌详情抽屉已打开。")
                try:
                    tab_res = await main_page.evaluate("""() => {
                        const drawer = document.querySelector('.okki-drawer.okki-drawer-open');
                        const btns = Array.from(drawer.querySelectorAll('.okki-tabs-tab-btn, .okki-tabs-tab'));
                        const target = btns.find(b => b.textContent.trim().includes('联系人'));
                        if (target) {
                            target.click();
                            return { ok: true, text: target.textContent.trim() };
                        }
                        return { ok: false };
                    }""")
                    log(f"  → 切换详情抽屉联系人 Tab 结果: {tab_res}")
                    await asyncio.sleep(3)
                    
                    company_contacts = await scrape_contacts_from_page(main_page, company, max_pages)
                except Exception as e:
                    log(f"  ❌ 抓取抽屉模式联系人出错: {e}")
                finally:
                    close_res = await main_page.evaluate("""() => {
                        const mask = document.querySelector('.okki-drawer-mask');
                        if (mask) {
                            mask.click();
                            return { ok: true };
                        }
                        return { ok: false };
                    }""")
                    log(f"  → 关闭详情抽屉: {close_res}")
                    await asyncio.sleep(2)
            else:
                log(f"  📋 [警告] 未检测到任何详情窗口或抽屉弹出，跳过。")
                continue
                
        log(f"  📋 [结果] ✅ 成功抓取到 {len(company_contacts)} 个真实联系人")
        all_contacts_database.extend(company_contacts)
        
        # Save temp backup
        if (idx + 1) % 10 == 0 or (idx + 1) == len(companies):
            backup_file = os.path.join(workspace_dir, "renwork_contacts_extractor_backup.json")
            with open(backup_file, "w", encoding="utf-8") as f:
                json.dump(all_contacts_database, f, ensure_ascii=False, indent=2)
            log(f"  💾 已安全保存阶段备份数据 ({len(all_contacts_database)} 条联系人)")

    # Export to sheet "17. 采购决策人联系方式"
    log("\n[RPA] 正在将所有提取的真实联系人数据写入 Excel Sheet...")
    if "17. 采购决策人联系方式" in wb.sheetnames:
        del wb["17. 采购决策人联系方式"]
        
    ws_out = wb.create_sheet(title="17. 采购决策人联系方式")
    headers_out = [
        "序号", "公司名称", "联系人姓名", "职务/Role", 
        "主邮箱", "备用邮箱/Other", "直拨电话/Mobile", "RenWork 匹配分数", "更新时间"
    ]
    ws_out.append(headers_out)
    
    for r_idx, c in enumerate(all_contacts_database):
        email_primary = c["emails"][0] if len(c["emails"]) > 0 else "N/A"
        email_secondary = c["emails"][1] if len(c["emails"]) > 1 else "N/A"
        phone = c["phones"][0] if len(c["phones"]) > 0 else "N/A"
        
        ws_out.append([
            r_idx + 1,
            c["company"],
            c["name"],
            c["title"],
            email_primary,
            email_secondary,
            phone,
            c["score"],
            c["date"]
        ])
        
    font_h = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_h = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for cell in ws_out[1]:
        cell.font = font_h
        cell.fill = fill_h
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    wb.save(excel_path)
    log(f"🎉 成功保存所有抓取数据至 {excel_path} 工作表 '17. 采购决策人联系方式'。")
    await browser.disconnect()

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="OKKI/RenWork Leads Contacts RPA Extractor")
    parser.add_argument("--excel", "-e", required=True, help="Path to Excel spreadsheet file")
    parser.add_argument("--sheet", "-s", default="", help="Name of sheet containing company names")
    parser.add_argument("--col", "-c", default="", help="Name of column containing company names")
    parser.add_argument("--pages", "-p", type=int, default=3, help="Maximum number of contact pagination pages to scrape per company")
    args = parser.parse_args()
    
    asyncio.run(run_leads_contacts_extraction_client(
        excel_path=args.excel,
        sheet_name=args.sheet,
        column_name=args.col,
        max_pages=args.pages
    ))
